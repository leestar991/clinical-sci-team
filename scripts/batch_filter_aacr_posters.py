#!/Volumes/data/github/clinical-sci-team/backend/.venv/bin/python
"""
批量处理 AACR Conference poster 筛选任务。

对 /Volumes/data/share/AACR-Confrence/tobe-summarize 下的每个 markdown 文件
和同名 excel 文件，调用 clinical-medicine agent 筛选最相关的 poster 记录，
输出带分类标签的新 excel 文件。

并发模型：2 个 worker，每个 worker 一次只处理一个任务。
worker 确认输出 excel 文件已写入 outputs 后，才从队列取下一个任务。
失败最多重试 5 次。按文件名升序排列处理。
"""

from __future__ import annotations

import asyncio
import argparse
import json
import logging
import shutil
import sys
import time
from pathlib import Path

import httpx
import openpyxl

# ── 配置 ─────────────────────────────────────────────────────────────────────
BASE_URL = "http://127.0.0.1:2026"
INPUT_DIR = Path("/Volumes/data/share/AACR-Confrence/tobe-summarize")
OUTPUT_DEST = Path("/Volumes/data/share/AACR-Confrence/excel-summarize")
POSTER_SRC = Path("/Volumes/data/share/AACR-Confrence/poster")
THREADS_BASE = Path("/Volumes/data/github/clinical-sci-team/backend/.deer-flow/threads")

PROMPT_TEMPLATE = (
    "# 任务指令（直接执行，禁止提问）\n\n"
    "你的唯一任务：从附件 Excel 中筛选与附件 Markdown 主题最相关的 poster，输出带分类标签的新 Excel。\n\n"
    "## 执行步骤\n\n"
    "步骤1：读取附件 Markdown 文件，提取核心主题、疾病领域、研究方向、关键词。\n\n"
    "步骤2：读取附件 Excel 文件，理解字段结构（Title、Abstract、match_score 等）。\n\n"
    "步骤3：筛选最相关的 poster（最多16条）。\n"
    "- 若有 match_score 字段，优先取高分记录，再结合语义相关性判断\n"
    "- 若无 match_score，纯按语义相关性判断\n\n"
    "步骤4：对筛选结果增加「分类标签」列，分为2-5个类别（如靶向治疗、免疫治疗、生物标志物、临床研究、基础研究等）。\n\n"
    "步骤5：将含分类标签的筛选结果写入 Excel 文件，保存路径：/mnt/user-data/outputs/{output_filename}\n\n"
    "步骤6：使用 present_files 工具展示输出文件路径。\n\n"
    "## 强制约束\n\n"
    "- 禁止追问、禁止请求确认、禁止列出候选方案\n"
    "- 禁止询问用户意见或偏好\n"
    "- 不要解释你的推理过程，直接输出结果\n"
    "- 输出文件名必须是：{output_filename}\n"
    "- 如果你不确定如何处理，选择最合理的默认方案直接执行\n"
)

NUM_WORKERS = 2          # 严格并发上限
MAX_RETRIES = 5          # 每个任务最多重试次数
TASK_TIMEOUT = 900       # 单次 agent 运行最长 15 分钟（关闭 plan_mode 后更快）
POST_STREAM_POLL = 180   # 流结束后再最多等 180 秒等文件落盘
POLL_INTERVAL = 5        # 轮询文件间隔（秒）

# ── 日志 ─────────────────────────────────────────────────────────────────────
LOG_FILE = "/tmp/batch_filter_aacr_posters.log"
_fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

_fh = logging.FileHandler(LOG_FILE, mode="a")
_fh.setFormatter(_fmt)

_sh = logging.StreamHandler(sys.stdout)
_sh.setFormatter(_fmt)

logging.root.setLevel(logging.INFO)
logging.root.addHandler(_fh)
if sys.stdout.isatty():
    logging.root.addHandler(_sh)
log = logging.getLogger(__name__)


def _tag(worker_id: int, task_idx: int) -> str:
    return f"[W{worker_id} T{task_idx:03d}]"


# ── API 调用 ──────────────────────────────────────────────────────────────────

async def create_thread(client: httpx.AsyncClient) -> str:
    """创建新 thread 并设置 agent_name 为 clinical-medicine。"""
    r = await client.post(f"{BASE_URL}/api/langgraph/threads", json={}, timeout=30)
    r.raise_for_status()
    thread_id: str = r.json()["thread_id"]
    await client.patch(
        f"{BASE_URL}/api/langgraph/threads/{thread_id}",
        json={"metadata": {"agent_name": "clinical-medicine"}},
        timeout=30,
    )
    return thread_id


async def upload_files(
    client: httpx.AsyncClient, thread_id: str, files: list[Path]
) -> list[dict]:
    """上传文件到 thread。"""
    handles = []
    for p in files:
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if p.suffix == ".xlsx" else "text/markdown"
        handles.append((p.name, open(p, "rb"), mime))
    try:
        r = await client.post(
            f"{BASE_URL}/api/threads/{thread_id}/uploads",
            files=[("files", h) for h in handles],
            timeout=120,
        )
        r.raise_for_status()
        return r.json()["files"]
    finally:
        for _, f, _ in handles:
            f.close()


async def stream_run(
    client: httpx.AsyncClient, thread_id: str, file_infos: list[dict], prompt: str
) -> None:
    """发送消息并完整读取 SSE 直到流关闭；收到 error 事件则抛出异常。"""
    files_meta = [
        {
            "filename": f["filename"],
            "size": str(f["size"]),
            "path": f["virtual_path"],
            "status": "uploaded",
        }
        for f in file_infos
    ]
    payload = {
        "assistant_id": "lead_agent",
        "input": {
            "messages": [
                {
                    "type": "human",
                    "content": [{"type": "text", "text": prompt}],
                    "additional_kwargs": {"files": files_meta},
                }
            ]
        },
        "config": {
            "recursion_limit": 1000,
            "configurable": {
                "agent_name": "clinical-medicine",
                "model_name": "qwen3-6-plus",
                "thinking_enabled": True,
                "reasoning_effort": "high",
                "is_plan_mode": True,         # 关闭 plan_mode 避免上下文摘要导致任务丢失
                "subagent_enabled": False,     # 单步任务无需子代理
                "thread_id": thread_id,
            },
        },
        "stream_mode": ["values", "messages", "custom"],
        "stream_subgraphs": True,
    }

    agent_error: str | None = None
    async with client.stream(
        "POST",
        f"{BASE_URL}/api/langgraph/threads/{thread_id}/runs/stream",
        json=payload,
        timeout=httpx.Timeout(TASK_TIMEOUT),
    ) as resp:
        resp.raise_for_status()
        async for line in resp.aiter_lines():
            if not line.startswith("data:"):
                continue
            raw = line[5:].strip()
            if not raw:
                continue
            try:
                obj = json.loads(raw)
                if isinstance(obj, dict) and "error" in obj:
                    agent_error = f"{obj['error']}: {obj.get('message', '')}"
            except json.JSONDecodeError:
                pass

    if agent_error:
        raise RuntimeError(f"Agent 返回错误: {agent_error}")


def find_output_excel(thread_id: str, expected_name: str) -> Path | None:
    """在 thread 的 outputs 目录查找输出的 excel 文件。"""
    out_dir = THREADS_BASE / thread_id / "user-data" / "outputs"
    if not out_dir.exists():
        return None
    # 精确匹配
    target = out_dir / expected_name
    if target.exists():
        return target
    # 模糊匹配：查找任何 xlsx 文件
    xlsx_files = list(out_dir.glob("*.xlsx"))
    if xlsx_files:
        return xlsx_files[0]
    return None


async def poll_until_file(thread_id: str, expected_name: str) -> Path | None:
    """流结束后再等最多 POST_STREAM_POLL 秒，直到输出 excel 文件落盘。"""
    deadline = time.monotonic() + POST_STREAM_POLL
    while time.monotonic() < deadline:
        result = find_output_excel(thread_id, expected_name)
        if result:
            return result
        await asyncio.sleep(POLL_INTERVAL)
    return find_output_excel(thread_id, expected_name)


# ── PDF 拷贝 ──────────────────────────────────────────────────────────────────

def _build_poster_index() -> dict[str, Path]:
    """遍历 POSTER_SRC 所有子目录，建立 {pdf文件名小写: 完整路径} 索引。"""
    index: dict[str, Path] = {}
    for pdf in POSTER_SRC.rglob("*.pdf"):
        index[pdf.name.lower()] = pdf
    log.info("Poster 索引构建完毕，共 %d 个 PDF 文件", len(index))
    return index


# 全局懒加载的 poster 索引（首次调用时构建）
_poster_index: dict[str, Path] | None = None


def get_poster_index() -> dict[str, Path]:
    global _poster_index
    if _poster_index is None:
        _poster_index = _build_poster_index()
    return _poster_index


def copy_posters_for_excel(tag: str, excel_path: Path, md_stem: str) -> int:
    """
    读取筛选结果 excel 的 'Poster File' 列，
    在 POSTER_SRC 中查找对应 PDF，复制到 OUTPUT_DEST/{md_stem}/ 目录。
    返回成功复制的文件数量。
    """
    index = get_poster_index()
    dest_dir = OUTPUT_DEST / md_stem
    dest_dir.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # 查找 Poster File 列（大小写不敏感，兼容中文"海报文件"）
        poster_col_idx = next(
            (i for i, h in enumerate(headers)
             if "poster file" in h.lower() or "poster_file" in h.lower() or "海报文件" in h),
            None,
        )
        if poster_col_idx is None:
            log.warning("%s Excel 中未找到 'Poster File' 列，跳过 PDF 拷贝。列名: %s", tag, headers)
            return 0

        copied = 0
        not_found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            pdf_name = row[poster_col_idx]
            if not pdf_name:
                continue
            pdf_name = str(pdf_name).strip()
            src = index.get(pdf_name.lower())
            if src:
                dest = dest_dir / src.name
                if not dest.exists():
                    shutil.copy2(src, dest)
                    log.info("%s ✓ PDF 已复制: %s → %s", tag, src.name, dest_dir)
                else:
                    log.info("%s ⏭ PDF 已存在，跳过: %s", tag, src.name)
                copied += 1
            else:
                not_found.append(pdf_name)

        if not_found:
            log.warning("%s %d 个 PDF 未找到: %s", tag, len(not_found), not_found)
        log.info("%s PDF 拷贝完成，共 %d 个文件 → %s", tag, copied, dest_dir)
        return copied

    except Exception as exc:
        log.error("%s PDF 拷贝失败: %s", tag, exc)
        return 0


# ── 单次尝试 ──────────────────────────────────────────────────────────────────

async def try_once(tag: str, md_file: Path, xlsx_file: Path, output_filename: str) -> bool:
    """尝试处理一对文件，成功返回 True。"""
    async with httpx.AsyncClient() as client:
        thread_id = await create_thread(client)
        log.info("%s Thread %s created", tag, thread_id)

        file_infos = await upload_files(client, thread_id, [md_file, xlsx_file])
        log.info("%s 已上传 %d 个文件: %s, %s", tag, len(file_infos), md_file.name, xlsx_file.name)

        prompt = PROMPT_TEMPLATE.format(output_filename=output_filename)

        t0 = time.monotonic()
        await stream_run(client, thread_id, file_infos, prompt)
        log.info("%s Agent 流结束，耗时 %.1fs", tag, time.monotonic() - t0)

    # 在流关闭后继续轮询文件
    output_file = await poll_until_file(thread_id, output_filename)

    if not output_file:
        log.warning("%s 未找到输出文件: %s", tag, output_filename)
        return False

    # 复制 Excel 到目标目录
    OUTPUT_DEST.mkdir(parents=True, exist_ok=True)
    dest_excel = OUTPUT_DEST / output_file.name
    shutil.copy2(output_file, dest_excel)
    log.info("%s ✓ Excel 已复制 %s → %s", tag, output_file.name, dest_excel)

    # 从 Excel 提取 Poster File 字段，拷贝对应 PDF
    copy_posters_for_excel(tag, dest_excel, md_file.stem)

    return True


# ── worker：从队列取任务，失败最多重试 MAX_RETRIES 次 ──────────────────────────

async def worker(worker_id: int, queue: asyncio.Queue, results: dict) -> None:
    while True:
        item = await queue.get()
        if item is None:  # 哨兵：退出
            queue.task_done()
            break

        task_idx, md_file, xlsx_file, output_filename = item
        tag = _tag(worker_id, task_idx)
        log.info("%s 开始处理: %s", tag, md_file.name)

        success = False
        for attempt in range(1, MAX_RETRIES + 1):
            log.info("%s 尝试 #%d/%d", tag, attempt, MAX_RETRIES)
            try:
                success = await try_once(tag, md_file, xlsx_file, output_filename)
            except Exception as exc:
                log.warning("%s 出错: %s", tag, exc)
                success = False

            if success:
                log.info("%s 任务完成（尝试 %d 次）", tag, attempt)
                results[md_file.stem] = "success"
                break

            if attempt < MAX_RETRIES:
                wait = min(15 * attempt, 120)
                log.info("%s 将在 %ds 后重试…", tag, wait)
                await asyncio.sleep(wait)

        if not success:
            log.error("%s 任务失败，已达最大重试次数 %d: %s", tag, MAX_RETRIES, md_file.name)
            results[md_file.stem] = "failed"

        queue.task_done()


# ── main ──────────────────────────────────────────────────────────────────────

def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="batch_filter_aacr_posters.py",
        description="批量筛选 AACR poster 并输出带分类标签的 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""示例:
  python batch_filter_aacr_posters.py                         # 处理所有未完成任务
  python batch_filter_aacr_posters.py -t T02                  # 仅处理 T02（前缀/子串匹配）
  python batch_filter_aacr_posters.py -t T02 --force          # 强制重跑，忽略已有输出
""",
    )
    parser.add_argument(
        "-t", "--task", metavar="TASK_NAME", default=None,
        help="仅处理此任务（精确名称或前缀/子串匹配）",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="强制重新处理，即使输出文件已存在",
    )
    return parser.parse_args()


async def main() -> None:
    args = _parse_args()

    # 收集所有需要处理的 md 文件（必须有同名 xlsx）
    all_md = sorted(INPUT_DIR.glob("*.md"))
    pairs: list[tuple[Path, Path]] = []
    for md in all_md:
        xlsx = md.with_suffix(".xlsx")
        if xlsx.exists():
            pairs.append((md, xlsx))
        else:
            log.warning("跳过 %s：找不到同名 xlsx 文件", md.name)

    # -t 参数：筛选匹配的任务（精确 or 前缀/子串）
    if args.task:
        matched = [
            (md, xlsx) for md, xlsx in pairs
            if md.stem == args.task
            or md.stem.startswith(args.task)
            or args.task in md.stem
        ]
        if not matched:
            log.error("找不到任务: %s（%s 中无匹配）", args.task, INPUT_DIR)
            sys.exit(1)
        if len(matched) > 1:
            log.warning("'%s' 匹配到多个任务，将全部处理: %s", args.task, [m.stem for m, _ in matched])
        pairs = matched
        log.info("单任务模式: %s", [m.stem for m, _ in pairs])

    # 检查已完成的文件（输出目录已有的）
    done_stems: set[str] = set()
    if not args.force and OUTPUT_DEST.exists():
        for f in OUTPUT_DEST.glob("*.xlsx"):
            # 从输出文件名反推源文件 stem
            # 输出命名格式: {stem}-filtered-posts.xlsx
            name = f.stem  # e.g. "01_IMG_xxx-filtered-posts"
            if name.endswith("-filtered-posts"):
                orig_stem = name[: -len("-filtered-posts")]
                done_stems.add(orig_stem)

    pending = [(md, xlsx) for md, xlsx in pairs if md.stem not in done_stems]

    log.info(
        "总文件: %d 对 | 已处理: %d | 待处理: %d | 并发: %d | 最大重试: %d",
        len(pairs), len(done_stems), len(pending), NUM_WORKERS, MAX_RETRIES,
    )

    if not pending:
        log.info("无待处理文件，退出。")
        return

    # 构建队列
    queue: asyncio.Queue = asyncio.Queue()
    results: dict[str, str] = {}

    for i, (md, xlsx) in enumerate(pending):
        output_filename = f"{md.stem}-filtered-posts.xlsx"
        queue.put_nowait((i + 1, md, xlsx, output_filename))

    # 为每个 worker 放一个哨兵
    for _ in range(NUM_WORKERS):
        queue.put_nowait(None)

    # 启动 worker
    workers = [
        asyncio.create_task(worker(w + 1, queue, results)) for w in range(NUM_WORKERS)
    ]
    await asyncio.gather(*workers)

    # 汇总
    success_count = sum(1 for v in results.values() if v == "success")
    fail_count = sum(1 for v in results.values() if v == "failed")
    total_done = len(list(OUTPUT_DEST.glob("*.xlsx"))) if OUTPUT_DEST.exists() else 0

    log.info("=" * 60)
    log.info("批处理完成")
    log.info("  本次成功: %d", success_count)
    log.info("  本次失败: %d", fail_count)
    log.info("  输出目录共: %d 个 xlsx 文件", total_done)
    log.info("=" * 60)

    if fail_count > 0:
        failed_files = [k for k, v in results.items() if v == "failed"]
        log.error("失败文件列表: %s", failed_files)
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
